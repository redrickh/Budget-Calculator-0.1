# Budget Calculator 0.1
# !/usr/bin/python
# -*- coding: utf-8 -*-

import openpyxl
import pandas as pd
import tkinter as tk
from tkinter import ttk, messagebox
from calendar import month_name
from tkinter import *
import matplotlib.pyplot as plt
import numpy as np
import os
from datetime import date

root = tk.Tk()
root.geometry('500x500')
root.resizable(False, False)
root.title('Budget Calculator 0.1')

label = ttk.Label(text="Please, select a month:")
label.pack(fill=tk.X, padx=5, pady=5)
selected_month = tk.StringVar()
month_cb = ttk.Combobox(root, textvariable=selected_month)
month_cb['values'] = [month_name[m][:3] for m in range(1, 13)]
month_cb['state'] = 'readonly'
month_cb.current(0)
month_cb.pack(fill=tk.X, padx=5, pady=5)

label = ttk.Label(text="Please, select a type:")
label.pack(fill=tk.X, padx=5, pady=5)
selected_title = tk.StringVar()
title_cb = ttk.Combobox(root, textvariable=selected_title)
title_cb["values"] = [""]
title_cb["state"] = "readonly"
title_cb.current(0)
title_cb.pack(fill=tk.X, padx=5, pady=5)

label = ttk.Label(text="Enter an amount:")
label.pack(fill=tk.X, padx=5, pady=5)
textbox = Text(root, height=1)
textbox.pack(fill=tk.X, padx=5, pady=5)

button = ttk.Button(root, text="Send")
button.pack(fill=tk.X, padx=5, pady=5)

button1 = ttk.Button(root, text="Statistics")
button1.pack(fill=tk.X, padx=5, pady=5)

listbox = Listbox(root, height=10,
                  width=15,
                  bg="grey",
                  activestyle='dotbox',
                  font="Helvetica",
                  fg="yellow")

listbox.pack(side=LEFT, padx=5, pady=5)

label = ttk.Label(text="Add a new type:")
label.pack(fill=tk.X, padx=5, pady=5)
textbox1 = Text(root, height=1)
textbox1.pack(fill=tk.X, padx=5, pady=5)

button2 = ttk.Button(root, text="Add")
button2.pack(fill=tk.X, padx=5, pady=5)

button3 = ttk.Button(root, text="Delete")
button3.pack(fill=tk.X, padx=5, pady=5)

labelY = ttk.Label(text="Year:")
labelY.pack(fill=tk.X, padx=5, pady=5)

selected_year = tk.StringVar()
year_cb = ttk.Combobox(root, textvariable=selected_year)
year_cb["values"] = [date.today().year]
year_cb["state"] = "readonly"
year_cb.current(0)
year_cb.pack(fill=tk.X, padx=5, pady=5)

buttonOpen = ttk.Button(root, text="Open my excel file")
buttonOpen.pack(fill=tk.X, padx=5, pady=5)

save_folder = os.getcwd() + "/income.xlsx"
print(save_folder)


class Load_Months:

    @classmethod
    def buildexcel(cls):
        if os.path.isfile(save_folder):
            print("Income.xlsx File exists")

        else:
            writer = pd.ExcelWriter(save_folder, engine='openpyxl')
            months = month_cb["values"]
            df = pd.DataFrame(list(months))
            df.to_excel(writer, index=False, sheet_name=str(selected_year.get()))
            months1 = title_cb["values"]
            df1 = pd.DataFrame(list(months1))
            df1 = df1.transpose()
            df1.to_excel(writer, sheet_name=str(selected_year.get()), startcol=1, index=False, header=False)
            writer.save()

    @classmethod
    def build_nums(cls):
        if os.path.isfile("nums.txt"):
            print("Income.xlsx File exists")
        else:
            r = open("nums.txt", "w")
            r.close()

    @classmethod
    def types(cls):
        months = month_cb["values"]
        months1 = title_cb["values"]

        workbook = openpyxl.load_workbook(save_folder)
        writer = pd.ExcelWriter(save_folder, engine='openpyxl')
        writer.book = workbook
        writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)
        df = pd.DataFrame(list(months))
        df1 = pd.DataFrame(list(months1))
        df1 = df1.transpose()
        df.to_excel(writer, sheet_name=str(selected_year.get()), index=False)
        df1.to_excel(writer, sheet_name=str(selected_year.get()), startcol=1, index=False, header=False)
        writer.save()
        writer.close()


class Main:
    Load_Months.build_nums()
    with open('nums.txt', 'r+') as file:
        for line in file:
            if not line.isspace():
                listbox.insert(END, line)

    title_cb["values"] = listbox.get(0, END)

    def add_listbox(self):
        if textbox1.compare("end-1c", "==", "1.0"):
            messagebox.showinfo("Empty data", "You cannot add empty text")

        else:
            listbox.insert(END, textbox1.get("1.0", 'end-1c'))
            title_cb["values"] = listbox.get(0, END)

            with open("nums.txt", "w") as Wr:
                for i in listbox.get(0, END):
                    Wr.write(i + "\n")

                    Load_Months.types()

    button2.bind('<Button-1>', add_listbox)

    def delete_listbox(self):
        curse = listbox.curselection()
        listbox.delete(curse)

        with open("nums.txt", "w") as op:
            title_cb["values"] = listbox.get(0, END)
            op.truncate(0)
            for i in listbox.get(0, END):
                op.write(i + "\n")

            addindex = list(curse)
            addindex_value = addindex[0]

            workbook = openpyxl.load_workbook(save_folder)
            writer = pd.ExcelWriter(save_folder, engine='openpyxl')
            writer.book = workbook

            df = pd.DataFrame([np.nan])
            writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)

            all_row = 0
            while all_row < 13:
                df.to_excel(writer, sheet_name=str(selected_year.get()), startcol=addindex_value + 1, startrow=all_row,
                            index=False, header=False)
                all_row += 1

            writer.save()
            writer.close()

    button3.bind('<Button-1>', delete_listbox)

    def enter_press(self):
        print("You hit return.")
        textbox.delete("1.0", END)
        textbox1.delete("1.0", END)
        messagebox.showinfo("Return", "You cannot use the Return key here.")

    def click(self):
        try:
            if textbox.compare("end-1c", "==", "1.0"):
                messagebox.showinfo("Empty data", "You cannot add empty text.")
            else:
                df = pd.DataFrame([int(textbox.get("1.0", 'end-1c'))])
                x = [month_name[m][:3] for m in range(1, 13)]
                y = title_cb["values"].index(selected_title.get())
                z = pd.read_excel(save_folder)
                sum_two_value = z.iloc[x.index(selected_month.get(), 0), y + 1]

                if np.isnan(sum_two_value):
                    print("cell value is empty")
                else:
                    df = df + sum_two_value

                workbook = openpyxl.load_workbook(save_folder)
                writer = pd.ExcelWriter(save_folder, engine='openpyxl')
                writer.book = workbook

                writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)
                df.to_excel(writer, sheet_name=str(selected_year.get()), startcol=y + 1,
                            startrow=x.index(selected_month.get(), 0) + 1,
                            index=False, header=False)
                writer.save()
                writer.close()

        except ValueError:
            title_cb.set("")
            print("ValueError: tuple.index(x): x not in tuple")
            messagebox.showinfo("Empty data", "You cannot add empty text.")

    button.bind('<Button-1>', click)
    root.bind("<Return>", enter_press)

    def graph(self):
        df = pd.read_excel(save_folder)

        df['sum'] = df[df.columns[1:]].sum(axis=1)
        # Thanks @Rabinzel for this solution
        # 'sum' will be new column
        # df1[df1.columns[1:]] will skip the first column
        # sum(axis=1) will add all values together along the axis (horizontally)
        # df1['sum'] is a pd.Series, if you want a list just do:
        result_list = df['sum'].tolist()

        plt.figure("Statistics")

        height = result_list
        bars = [month_name[m][:3] for m in range(1, 13)]
        y_pos = np.arange(len(bars))

        plt.bar(y_pos, height)

        plt.xticks(y_pos, bars)

        plt.show()

    def open_my_excel(self):
        try:
            os.system(save_folder)
        except Exception as e:
            print(e)

    button1.bind('<Button-1>', graph)
    buttonOpen.bind('<Button-1>', open_my_excel)


Load_Months.build_nums()
Load_Months.buildexcel()
Load_Months.types()

root.mainloop()
